from django.shortcuts import render
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from shop.models import Order, OrderItem
from django.utils import timezone
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status
# Add this import at the top
from datetime import datetime, timedelta
from import_export.formats.base_formats import XLSX
from .resources import TransactionResource
from django.http import HttpResponse
from datetime import datetime, timedelta
import pytz
import json
from managers.decorators import manager_required
from managepayments.models import DeliveryInfo, DeliveryStatus


@manager_required
def transactions_view(request):
    """View for displaying the transactions page"""
    return render(request, 'transactions/transactions.html')


@manager_required
def get_transactions(request):
    """API endpoint to get all transactions"""
    try:
        # Get all orders
        orders = Order.objects.all().order_by('-date_created')
        
        # Format the orders
        formatted_orders = []
        for order in orders:
            # Calculate total
            total = sum(item.price * item.quantity for item in OrderItem.objects.filter(order=order))
            
            # Format the date
            date_timestamp = int(order.date_created.timestamp() * 1000)
            
            # Format the order
            formatted_orders.append({
                'order_id': order.order_id,
                'student_id': order.student_id,
                'date': date_timestamp,
                'total': float(total),
                'status': order.status  # Make sure this is included!
            })
        
        return JsonResponse({
            'success': True,
            'transactions': formatted_orders
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@manager_required
def get_order_details(request, order_id):
    """API endpoint to get details of a specific order"""
    try:
        # Get the order
        order = Order.objects.get(order_id=order_id)
        
        # Get order items
        order_items = OrderItem.objects.filter(order=order)
        
        # Calculate total
        total = sum(item.price * item.quantity for item in order_items)
        
        # Format the date
        date_timestamp = int(order.date_created.timestamp() * 1000)
        
        # Format the items
        items = [
            {
                'name': item.name,
                'price': float(item.price),
                'quantity': item.quantity
            } for item in order_items
        ]
        
        # Format the order data
        order_data = {
            'order_id': order.order_id,
            'student_id': order.student_id,
            'date': date_timestamp,
            'total': float(total),
            'items': items,
            'name': order.name,
            'payment_method': order.payment_method,
            'status': order.status  # Include status in response
        }
        
        return JsonResponse({
            'success': True,
            'order': order_data
        })
    except Order.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': f'Order with ID {order_id} not found'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)
    


# Add this new function to the file
# @manager_required
# def export_transactions(request):
#     """Export transactions to Excel based on filter with statistics"""
#     from openpyxl import Workbook
#     from openpyxl.styles import Font, Alignment, PatternFill
#     # Add these imports for charts
#     from openpyxl.chart import PieChart, Reference
#     from openpyxl.chart.label import DataLabelList
    
#     filter_type = request.GET.get('filter', 'all')
    
#     # Apply the right filter based on selection
#     now = datetime.now()
#     if filter_type == 'today':
#         today = now.date()
#         queryset = Order.objects.filter(date_created__date=today)
#         filename = f"transactions_{today}.xlsx"
#         date_range = f"Date: {today}"
#     elif filter_type == 'week':
#         week_ago = now - timedelta(days=7)
#         queryset = Order.objects.filter(date_created__gte=week_ago)
#         filename = "transactions_last7days.xlsx"
#         date_range = f"Period: {week_ago.date()} to {now.date()}"
#     elif filter_type == 'month':
#         month_ago = now - timedelta(days=30)
#         queryset = Order.objects.filter(date_created__gte=month_ago)
#         filename = "transactions_last30days.xlsx"
#         date_range = f"Period: {month_ago.date()} to {now.date()}"
#     else:
#         queryset = Order.objects.all()
#         filename = "all_transactions.xlsx"
#         date_range = "All Time"
    
#     # Create workbook and sheets
#     wb = Workbook()
#     stats_sheet = wb.active
#     stats_sheet.title = "Summary Statistics"
#     data_sheet = wb.create_sheet(title="Transaction Data")
    
#     # Calculate statistics for the report
#     total_transactions = queryset.count()
#     total_revenue = sum(sum(item.price * item.quantity for item in OrderItem.objects.filter(order=order)) for order in queryset) if total_transactions > 0 else 0
#     avg_order_value = total_revenue / total_transactions if total_transactions > 0 else 0
    
#     # Get the items statistics
#     item_counts = {}
#     order_items = OrderItem.objects.filter(order__in=queryset)
#     for item in order_items:
#         if item.name in item_counts:
#             item_counts[item.name] += item.quantity
#         else:
#             item_counts[item.name] = item.quantity
    
#     sorted_items = sorted(item_counts.items(), key=lambda x: x[1], reverse=True)
#     top_items = sorted_items[:5] if len(sorted_items) >= 5 else sorted_items
#     least_items = sorted_items[-5:] if len(sorted_items) >= 5 else sorted_items
#     least_items.reverse()
    
#     # Create header style
#     header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
#     header_font = Font(color="FFFFFF", bold=True)
#     header_align = Alignment(horizontal="center", vertical="center")
    
#     # Add statistics to summary sheet
#     stats_sheet['A1'] = "Cafeteria Management System - Transaction Report"
#     stats_sheet['A1'].font = Font(bold=True, size=16)
#     stats_sheet.merge_cells('A1:D1')
    
#     stats_sheet['A3'] = date_range
#     stats_sheet['A3'].font = Font(bold=True)

#     # Add download timestamp (using IST timezone)
#     ist_timezone = pytz.timezone('Asia/Kolkata')
#     ist_now = now.astimezone(ist_timezone)
#     download_time = ist_now.strftime('%d/%m/%Y %H:%M:%S')
#     stats_sheet['A4'] = f"Report Downloaded On: {download_time} IST"
#     stats_sheet['A4'].font = Font(italic=True)
    
#     # Add summary statistics
#     stats_sheet['A6'] = "Total Transactions:"
#     stats_sheet['B6'] = total_transactions
    
#     stats_sheet['A7'] = "Total Revenue:"
#     stats_sheet['B7'] = f"₹{total_revenue:.2f}"
    
#     stats_sheet['A8'] = "Average Order Value:"
#     stats_sheet['B8'] = f"₹{avg_order_value:.2f}"
    
#     # Add top items
#     stats_sheet['A10'] = "Top Sold Items"
#     stats_sheet['A10'].font = Font(bold=True)
#     stats_sheet['A11'] = "Item Name"
#     stats_sheet['B11'] = "Quantity"
    
#     # Header style for item tables
#     for cell in stats_sheet['A11:B11'][0]:
#         cell.font = Font(bold=True)
#         cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    
#     # Add top items data
#     row = 12
#     for item_name, count in top_items:
#         stats_sheet[f'A{row}'] = item_name
#         stats_sheet[f'B{row}'] = count
#         row += 1
    
#     top_items_end_row = row - 1
    
#     # Create pie chart for top items
#     pie1 = PieChart()
#     pie1.title = "Top Sold Items"
#     labels = Reference(stats_sheet, min_col=1, min_row=12, max_row=top_items_end_row)
#     data = Reference(stats_sheet, min_col=2, min_row=11, max_row=top_items_end_row)
#     pie1.add_data(data, titles_from_data=True)
#     pie1.set_categories(labels)
#     pie1.height = 10
#     pie1.width = 15
    
#     # Add data labels showing percentages
#     pie1.dataLabels = DataLabelList()
#     pie1.dataLabels.showPercent = True
#     pie1.dataLabels.showVal = True
#     pie1.dataLabels.showCatName = True
    
#     # Add the chart to sheet
#     stats_sheet.add_chart(pie1, "E15")
    
#     # Add least sold items with a gap
#     stats_sheet[f'A{row+2}'] = "Least Sold Items"
#     stats_sheet[f'A{row+2}'].font = Font(bold=True)
#     stats_sheet[f'A{row+3}'] = "Item Name"
#     stats_sheet[f'B{row+3}'] = "Quantity"
    
#     # Header style for least items table
#     for cell in stats_sheet[f'A{row+3}:B{row+3}'][0]:
#         cell.font = Font(bold=True)
#         cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    
#     least_start_row = row + 4
#     row = least_start_row
    
#     for item_name, count in least_items:
#         stats_sheet[f'A{row}'] = item_name
#         stats_sheet[f'B{row}'] = count
#         row += 1
    
#     least_end_row = row - 1
    
#     # Create pie chart for least sold items
#     pie2 = PieChart()
#     pie2.title = "Least Sold Items"
#     labels = Reference(stats_sheet, min_col=1, min_row=least_start_row, max_row=least_end_row)
#     data = Reference(stats_sheet, min_col=2, min_row=least_start_row-1, max_row=least_end_row)
#     pie2.add_data(data, titles_from_data=True)
#     pie2.set_categories(labels)
#     pie2.height = 10
#     pie2.width = 15
    
#     # Add data labels showing percentages
#     pie2.dataLabels = DataLabelList()
#     pie2.dataLabels.showPercent = True
#     pie2.dataLabels.showVal = False
#     pie2.dataLabels.showCatName = True
    
#     # Add the chart to sheet - position it below the first chart
#     stats_sheet.add_chart(pie2, "E37")
    
#     # Set column widths
#     for col in ['A', 'B', 'C', 'D']:
#         stats_sheet.column_dimensions[col].width = 20
    

#     stats_sheet['A8'] = "Average Order Value:"
#     stats_sheet['B8'] = f"₹{avg_order_value:.2f}"








#     # Calculate payment method statistics
#     cash_orders = 0
#     online_orders = 0
#     deliver_to_class = 0

#     for order in queryset:
#         # Check payment method
#         payment_method = order.payment_method.lower() if hasattr(order, 'payment_method') and order.payment_method else ""
        
#         # Count by payment method
#         if payment_method in ['cash', 'cod', 'cash on delivery']:
#             cash_orders += 1
#         elif 'upi' in payment_method or payment_method == 'online':
#             online_orders += 1
            
#         # Check for delivery info - independent of payment method
#         if hasattr(order, 'delivery_info') and order.delivery_info:
#             deliver_to_class += 1

#     # Calculate percentages
#     cash_percentage = (cash_orders / total_transactions * 100) if total_transactions > 0 else 0
#     online_percentage = (online_orders / total_transactions * 100) if total_transactions > 0 else 0
#     delivery_percentage = (deliver_to_class / total_transactions * 100) if total_transactions > 0 else 0

#     # Add payment method breakdown as a separate section - don't change existing row structure
#     stats_sheet['D6'] = "Payment Method Breakdown"
#     stats_sheet['D6'].font = Font(bold=True)

#     # Add header row
#     stats_sheet['D7'] = "Payment Type"
#     stats_sheet['E7'] = "Count"
#     stats_sheet['F7'] = "Percentage"

#     # Style the header row
#     for cell in stats_sheet['D7:F7'][0]:
#         cell.font = Font(bold=True)
#         cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

#     # Add the data rows
#     stats_sheet['D8'] = "Cash Orders"
#     stats_sheet['E8'] = cash_orders
#     stats_sheet['F8'] = f"{cash_percentage:.1f}%"

#     stats_sheet['D9'] = "Online/UPI Orders"
#     stats_sheet['E9'] = online_orders
#     stats_sheet['F9'] = f"{online_percentage:.1f}%"

#     stats_sheet['D10'] = "Deliver to Class"
#     stats_sheet['E10'] = deliver_to_class
#     stats_sheet['F10'] = f"{delivery_percentage:.1f}%"

#     # Add top items section - using the original row numbers
#     stats_sheet['A10'] = "Top Sold Items"
#     stats_sheet['A10'].font = Font(bold=True)
#     stats_sheet['A11'] = "Item Name"
#     stats_sheet['B11'] = "Quantity"
#     # order id numbers for all the items






    



    
#     # Now add transaction data to data sheet
#     # Add headers
#     headers = ['Order ID', 'Student ID', 'Name', 'Date', 'Total Amount', 'Payment Method', 'Status']
#     for col_idx, header in enumerate(headers, 1):
#         cell = data_sheet.cell(row=1, column=col_idx)
#         cell.value = header
#         cell.fill = header_fill
#         cell.font = header_font
#         cell.alignment = header_align
        
#     # Add transaction data
#     row = 2
#     for order in queryset:
#         # Calculate total for this order
#         order_total = sum(item.price * item.quantity for item in OrderItem.objects.filter(order=order))
        
#         data_sheet.cell(row=row, column=1).value = order.order_id
#         data_sheet.cell(row=row, column=2).value = order.student_id
#         data_sheet.cell(row=row, column=3).value = order.name if hasattr(order, 'name') and order.name else ''
#         data_sheet.cell(row=row, column=4).value = order.date_created.strftime('%Y-%m-%d %H:%M:%S')
#         data_sheet.cell(row=row, column=5).value = float(order_total)
#         data_sheet.cell(row=row, column=6).value = order.payment_method if hasattr(order, 'payment_method') and order.payment_method else ''
#         data_sheet.cell(row=row, column=7).value = order.status.title() if hasattr(order, 'status') and order.status else 'Pending'
#         row += 1
    
#     # Set column widths for data sheet
#     for col_idx, width in enumerate([15, 15, 20, 20, 15, 20, 15], 1):
#         data_sheet.column_dimensions[chr(64 + col_idx)].width = width
    

    
    
#     # Create the HTTP response with Excel content
#     response = HttpResponse(
#         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#     )
#     response['Content-Disposition'] = f'attachment; filename="{filename}"'
#     wb.save(response)
    
#     return response


@manager_required
def export_transactions(request):
    """Export transactions to Excel based on filter with statistics"""
    # --- IMPORT LIBRARIES ---
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    # Add these imports for charts
    from openpyxl.chart import PieChart, Reference
    from openpyxl.chart.label import DataLabelList
    
    # --- FILTER SETUP ---
    # Get filter type parameter from request (defaults to 'all')
    filter_type = request.GET.get('filter', 'all')
    
    # --- TRANSACTION FILTERING BASED ON TIME PERIOD ---
    now = datetime.now()
    if filter_type == 'today':
        today = now.date()
        queryset = Order.objects.filter(date_created__date=today)
        filename = f"transactions_{today}.xlsx"
        date_range = f"Date: {today}"
    elif filter_type == 'week':
        week_ago = now - timedelta(days=7)
        queryset = Order.objects.filter(date_created__gte=week_ago)
        filename = "transactions_last7days.xlsx"
        date_range = f"Period: {week_ago.date()} to {now.date()}"
    elif filter_type == 'month':
        month_ago = now - timedelta(days=30)
        queryset = Order.objects.filter(date_created__gte=month_ago)
        filename = "transactions_last30days.xlsx"
        date_range = f"Period: {month_ago.date()} to {now.date()}"
    else:
        # Default: all transactions
        queryset = Order.objects.all()
        filename = "all_transactions.xlsx"
        date_range = "All Time"
    
    # --- EXCEL WORKBOOK CREATION ---
    # Create workbook with two sheets: stats and transaction data
    wb = Workbook()
    stats_sheet = wb.active
    stats_sheet.title = "Summary Statistics"
    data_sheet = wb.create_sheet(title="Transaction Data")
    
    # --- CALCULATE BASIC STATISTICS ---
    # Count and financial calculations
    total_transactions = queryset.count()
    total_revenue = sum(sum(item.price * item.quantity for item in OrderItem.objects.filter(order=order)) for order in queryset) if total_transactions > 0 else 0
    avg_order_value = total_revenue / total_transactions if total_transactions > 0 else 0
    
    # --- CALCULATE ITEM STATISTICS ---
    # Build dictionary of item counts from order items
    item_counts = {}
    order_items = OrderItem.objects.filter(order__in=queryset)
    for item in order_items:
        if item.name in item_counts:
            item_counts[item.name] += item.quantity
        else:
            item_counts[item.name] = item.quantity
    
    # Sort items by popularity for charts
    sorted_items = sorted(item_counts.items(), key=lambda x: x[1], reverse=True)
    top_items = sorted_items[:5] if len(sorted_items) >= 5 else sorted_items
    least_items = sorted_items[-5:] if len(sorted_items) >= 5 else sorted_items
    least_items.reverse()
    
    # --- STYLE DEFINITIONS ---
    # Define cell styles for consistent formatting
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_align = Alignment(horizontal="center", vertical="center")
    
    # --- REPORT HEADER SECTION ---
    # Add report title
    stats_sheet['A1'] = "Cafeteria Management System - Transaction Report"
    stats_sheet['A1'].font = Font(bold=True, size=16)
    stats_sheet.merge_cells('A1:D1')
    
    # Add date range information
    stats_sheet['A3'] = date_range
    stats_sheet['A3'].font = Font(bold=True)

    # Add download timestamp in IST timezone
    ist_timezone = pytz.timezone('Asia/Kolkata')
    ist_now = now.astimezone(ist_timezone)
    download_time = ist_now.strftime('%d/%m/%Y %H:%M:%S')
    stats_sheet['A4'] = f"Report Downloaded On: {download_time} IST"
    stats_sheet['A4'].font = Font(italic=True)
    
    # --- BASIC STATISTICS SECTION ---
    # Add key metrics to the report
    stats_sheet['A6'] = "Total Transactions:"
    stats_sheet['B6'] = total_transactions
    
    stats_sheet['A7'] = "Total Revenue:"
    stats_sheet['B7'] = f"₹{total_revenue:.2f}"
    
    stats_sheet['A8'] = "Average Order Value:"
    stats_sheet['B8'] = f"₹{avg_order_value:.2f}"
    
    # --- TOP SOLD ITEMS SECTION ---
    # Add section title and headers
    stats_sheet['A10'] = "Top Sold Items"
    stats_sheet['A10'].font = Font(bold=True)
    stats_sheet['A11'] = "Item Name"
    stats_sheet['B11'] = "Quantity"
    
    # Style the header row
    for cell in stats_sheet['A11:B11'][0]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    
    # Add top selling items data to the table
    row = 12
    for item_name, count in top_items:
        stats_sheet[f'A{row}'] = item_name
        stats_sheet[f'B{row}'] = count
        row += 1
    
    top_items_end_row = row - 1
    
    # --- TOP ITEMS CHART ---
    # Create pie chart for top selling items
    pie1 = PieChart()
    pie1.title = "Top Sold Items"
    labels = Reference(stats_sheet, min_col=1, min_row=12, max_row=top_items_end_row)
    data = Reference(stats_sheet, min_col=2, min_row=11, max_row=top_items_end_row)
    pie1.add_data(data, titles_from_data=True)
    pie1.set_categories(labels)
    pie1.height = 10
    pie1.width = 15
    
    # Configure chart data labels
    pie1.dataLabels = DataLabelList()
    pie1.dataLabels.showPercent = True
    pie1.dataLabels.showVal = True
    pie1.dataLabels.showCatName = True
    
    # Position the chart in the sheet
    stats_sheet.add_chart(pie1, "E15")
    
    # --- LEAST SOLD ITEMS SECTION ---
    # Add section title with spacing after top items
    stats_sheet[f'A{row+2}'] = "Least Sold Items"
    stats_sheet[f'A{row+2}'].font = Font(bold=True)
    stats_sheet[f'A{row+3}'] = "Item Name"
    stats_sheet[f'B{row+3}'] = "Quantity"
    
    # Style the header row
    for cell in stats_sheet[f'A{row+3}:B{row+3}'][0]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    
    # Add least selling items data to the table
    least_start_row = row + 4
    row = least_start_row
    
    for item_name, count in least_items:
        stats_sheet[f'A{row}'] = item_name
        stats_sheet[f'B{row}'] = count
        row += 1
    
    least_end_row = row - 1
    
    # --- LEAST ITEMS CHART ---
    # Create pie chart for least selling items
    pie2 = PieChart()
    pie2.title = "Least Sold Items"
    labels = Reference(stats_sheet, min_col=1, min_row=least_start_row, max_row=least_end_row)
    data = Reference(stats_sheet, min_col=2, min_row=least_start_row-1, max_row=least_end_row)
    pie2.add_data(data, titles_from_data=True)
    pie2.set_categories(labels)
    pie2.height = 10
    pie2.width = 15
    
    # Configure chart data labels
    pie2.dataLabels = DataLabelList()
    pie2.dataLabels.showPercent = True
    pie2.dataLabels.showVal = False
    pie2.dataLabels.showCatName = True
    
    # Position the chart below the first chart
    stats_sheet.add_chart(pie2, "E37")
    
    # Set column widths for better readability
    for col in ['A', 'B', 'C', 'D']:
        stats_sheet.column_dimensions[col].width = 20
    
    # --- DUPLICATE AVERAGE ORDER VALUE (redundant, can be removed) ---
    stats_sheet['A8'] = "Average Order Value:"
    stats_sheet['B8'] = f"₹{avg_order_value:.2f}"

    # --- PAYMENT METHOD STATISTICS SECTION ---
    # Initialize counters for payment methods
    cash_orders = 0
    online_orders = 0
    deliver_to_class = 0

    # Count orders by payment method and delivery type
    for order in queryset:
        # Check payment method with safety check for attribute existence
        payment_method = order.payment_method.lower() if hasattr(order, 'payment_method') and order.payment_method else ""
        
        # Classify and count by payment method
        if payment_method in ['cash', 'cod', 'cash on delivery']:
            cash_orders += 1
        elif 'upi' in payment_method or payment_method == 'online':
            online_orders += 1
            
        # Count orders with delivery info (class delivery) - independent of payment method
        if hasattr(order, 'delivery_info') and order.delivery_info:
            deliver_to_class += 1
        # Also check if payment_method contains classroom_delivery
        elif payment_method and 'classroom_delivery' in payment_method:
            deliver_to_class += 1

    # Calculate percentages with division by zero protection
    cash_percentage = (cash_orders / total_transactions * 100) if total_transactions > 0 else 0
    online_percentage = (online_orders / total_transactions * 100) if total_transactions > 0 else 0
    delivery_percentage = (deliver_to_class / total_transactions * 100) if total_transactions > 0 else 0

    # Add payment method breakdown table header
    stats_sheet['D6'] = "Payment Method Breakdown"
    stats_sheet['D6'].font = Font(bold=True)

    # Add header row for payment method table
    stats_sheet['D7'] = "Payment Type"
    stats_sheet['E7'] = "Count"
    stats_sheet['F7'] = "Percentage"

    # Style the header row
    for cell in stats_sheet['D7:F7'][0]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

    # Add payment method statistics data rows
    stats_sheet['D8'] = "Cash Orders"
    stats_sheet['E8'] = cash_orders
    stats_sheet['F8'] = f"{cash_percentage:.1f}%"

    stats_sheet['D9'] = "Online/UPI Orders"
    stats_sheet['E9'] = online_orders
    stats_sheet['F9'] = f"{online_percentage:.1f}%"

    stats_sheet['D10'] = "Deliver to Class"
    stats_sheet['E10'] = deliver_to_class
    stats_sheet['F10'] = f"{delivery_percentage:.1f}%"

    # --- REDUNDANT TOP ITEMS SECTION HEADERS (duplicated code, can be removed) ---
    stats_sheet['A10'] = "Top Sold Items"
    stats_sheet['A10'].font = Font(bold=True)
    stats_sheet['A11'] = "Item Name"
    stats_sheet['B11'] = "Quantity"
    
    # --- TRANSACTION DATA SHEET ---
    # Define columns for transaction details
    headers = ['Order ID', 'Student ID', 'Name', 'Date', 'Total Amount', 'Payment Method', 'Status']
    for col_idx, header in enumerate(headers, 1):
        cell = data_sheet.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        
    # Add individual transaction rows to the data sheet
    row = 2
    for order in queryset:
        # Calculate total amount for this specific order
        order_total = sum(item.price * item.quantity for item in OrderItem.objects.filter(order=order))
        
        # Fill in transaction details with safety checks for missing attributes
        data_sheet.cell(row=row, column=1).value = order.order_id
        data_sheet.cell(row=row, column=2).value = order.student_id
        data_sheet.cell(row=row, column=3).value = order.name if hasattr(order, 'name') and order.name else ''
        data_sheet.cell(row=row, column=4).value = order.date_created.strftime('%Y-%m-%d %H:%M:%S')
        data_sheet.cell(row=row, column=5).value = float(order_total)
        data_sheet.cell(row=row, column=6).value = order.payment_method if hasattr(order, 'payment_method') and order.payment_method else ''
        data_sheet.cell(row=row, column=7).value = order.status.title() if hasattr(order, 'status') and order.status else 'Pending'
        row += 1
    
    # Set column widths for better readability in data sheet
    for col_idx, width in enumerate([15, 15, 20, 20, 15, 20, 15], 1):
        data_sheet.column_dimensions[chr(64 + col_idx)].width = width
    
    # --- GENERATE AND RETURN EXCEL RESPONSE ---
    # Create HTTP response with Excel file
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    
    return response



@manager_required
def update_order_status(request, order_id):
    """API endpoint to update order status"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Only POST method is allowed'}, status=405)
    
    try:
        # Parse request data
        data = json.loads(request.body)
        new_status = data.get('status')
        delivery_id = data.get('delivery_id')
        
        # Validate status
        valid_statuses = ['pending', 'in_progress', 'successful', 'cancelled']
        if not new_status or new_status not in valid_statuses:
            return JsonResponse({
                'success': False,
                'error': 'Invalid status value'
            }, status=400)
        
        # Update order status
        order = Order.objects.get(order_id=order_id)
        order.status = new_status
        order.save()
        
        # IMPORTANT - Update the delivery status directly
        if delivery_id:
            try:
                # Use get() not filter() to ensure we get the right record
                delivery = DeliveryInfo.objects.get(id=delivery_id)
                
                # Important change: Map successful status to "delivered"
                # and make sure to use the right status field name
                delivery_status = "delivered" if new_status == "successful" else new_status
                
                # Debug output before update
                print(f"Before update - Delivery {delivery_id} status: {delivery.status}")
                
                # Update status and save
                delivery.status = delivery_status
                delivery.save()
                
                # Verify the update worked
                refreshed = DeliveryInfo.objects.get(id=delivery_id)
                print(f"After update - Delivery {delivery_id} status: {refreshed.status}")
                
                # Return the updated status in the response
                return JsonResponse({
                    'success': True,
                    'order_id': order_id,
                    'new_status': new_status,
                    'delivery_status': delivery_status
                })
            
            except Exception as e:
                print(f"ERROR updating delivery status: {str(e)}")
                # Still return success for the order update
                return JsonResponse({
                    'success': True,
                    'order_id': order_id,
                    'new_status': new_status,
                    'error': f"Delivery update failed: {str(e)}"
                })
        
        # If no delivery_id provided
        return JsonResponse({
            'success': True,
            'order_id': order_id,
            'new_status': new_status
        })
        
    except Order.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': f'Order with ID {order_id} not found'
        }, status=404)
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Invalid JSON in request body'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)











@manager_required
def view_delivery_orders(request):
    """View for displaying classroom delivery orders from DeliveryInfo model"""
    # Get all delivery info records, most recent first
    delivery_orders = DeliveryInfo.objects.select_related('order').order_by('-order__date')
    
    context = {
        'delivery_orders': delivery_orders
    }
    return render(request, 'transactions/view_deliver_to_class.html', context)







